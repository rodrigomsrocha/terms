from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment

filename = "hello.xlsx"

workbook = load_workbook(filename)
worksheet = workbook.active

def getLastIndex(value):
  lastIndex = None
  for row in worksheet.iter_rows():
    for cell in row:
      if cell.value == value:
        lastIndex = cell.row
  return lastIndex

while True:
  code = input("nº termo:")
  day = int(input("dia:"))
  month = int(input("mês:"))
  data = datetime(2022, month, day)
  
  newTermRow = getLastIndex(data) + 1
  worksheet["A{row}".format(row = newTermRow)] = code
  worksheet["B{row}".format(row = newTermRow)] = data

  launchAnother = input("Lançar outro? s/n:")
  if launchAnother != "s":
    break

for cell in worksheet["A"]:
  cell.alignment = Alignment(horizontal= "right")

workbook.save(filename)
